import os
import shutil
import subprocess
import tkinter as tk
from datetime import datetime
from tkinter import (ttk, messagebox, filedialog)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from database import (check_song_exists, insert_into_songs, insert_into_error_log, get_all_songs,
                      get_songs_by_name, get_all_error_logs, get_processing_error_logs, delete_song, delete_error_log,
                      database_path)
from metadata_extractor import (extract_metadata, is_valid_folder)
from constants import (SONG_SEARCH_BAR_PLACEHOLDER)
from settings import (IS_TEST_MODE, APP_ROOT_FOLDER_TEST_MODE, APP_ROOT_FOLDER, APPROVED_MUSIC_FOLDER, DATABASE_FILE_NAME)
from enums import (ErrorType)


# ---- Treeview Setup ----
def setup_treeview(parent, column_properties):
    try:
        columns = list(column_properties.keys())  # extract column names from the keys of column_properties

        tree = ttk.Treeview(parent, columns=columns, show="headings", height=10)
        tree.pack(padx=10, pady=10, fill='both', expand=True)

        for col, properties in column_properties.items():
            # Set the heading and apply width and alignment
            #   -The 'get' function looks for the specific key, if it cannot find it, it uses the default
            #   -E.G. properties.get("width", 120) - looks for "width", if not found uses 120.
            tree.heading(col, text=col)
            tree.column(col, width=properties.get("width", 120), anchor=properties.get("anchor", "center"))

        return tree

    except Exception as e:
        # Log the unexpected error
        error_id = insert_into_error_log(ErrorType.UNKNOWN_ERROR,
                                         f"Unexpected error in setup_treeview: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred while setting up the treeview. " +
                             f"See Error Log - Error ID {error_id}."
                             )


# ---- Custom Button Styles ----
def create_custom_style(root):
    try:
        # Create and apply a custom style for ttk buttons.
        style = ttk.Style(root)
        button_font = ("Arial", 9, "bold")

        # Default background color and font
        style.configure("Custom.TButton", font=button_font, background="#555555", foreground="#FFFFFF")

        # Different shades for different button states
        style.map("Custom.TButton",
                  background=[("active", "#FFFFFF"),  # hover effect
                              ("pressed", "#FFFFFF"),  # clicked effect
                              ("!active", "#555555")],  # default state
                  foreground=[("disabled", "#AAAAAA")])  # gray out text if button is disabled

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.STYLING_ERROR,
                                         f"Unexpected error in create_custom_style: {e}"
                                         )
        messagebox.showerror("Error",
                             f"An unexpected error occurred while setting up the button style. "
                             f"See Error Log - Error ID {error_id}."
                             )


def create_button(parent, text, command):
    try:
        # Helper function to create a styled button with predefined style.
        return ttk.Button(parent, text=text, style="Custom.TButton", command=command)

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.STYLING_ERROR,
                                         f"Unexpected error in create_button: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred while creating a button. "
                             f"See Error Log - Error ID {error_id}."
                             )


# ---- Search Bar (Database Window) ----
def set_search_bar_placeholder(search_bar):
    try:
        # Set the search bar placeholder text if the search_bar is empty.
        if search_bar.get() == "":
            search_bar.insert(0, SONG_SEARCH_BAR_PLACEHOLDER)
            search_bar.configure(fg="gray")

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.STYLING_ERROR,
                                         f"Unexpected error in set_search_bar_placeholder: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred while setting the search bar placeholder. " +
                             f"See Error Log - Error ID {error_id}."
                             )


def on_focus_in(event):
    try:
        # Remove the placeholder text when search_bar is focused.
        search_bar = event.widget
        if search_bar.get() == SONG_SEARCH_BAR_PLACEHOLDER:
            search_bar.delete(0, tk.END)
            search_bar.configure(fg="white")

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.EVENT_ERROR,
                                         f"Unexpected error in on_focus_in: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred while focusing on the search bar. "
                             f"See Error Log - Error ID {error_id}."
                             )


def on_focus_out(event):
    try:
        # Set the placeholder text if search_bar is empty on focus out.
        search_bar = event.widget
        set_search_bar_placeholder(search_bar)

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.EVENT_ERROR,
                                         f"Unexpected error in on_focus_out: {e}"
                                         )
        messagebox.showerror("Error",
                             f"An unexpected error occurred when focus left the search bar. "
                             f"See Error Log - Error ID {error_id}."
                             )


# ---- Data Processing ----
def on_drop(event, tree, status_font):
    try:
        file_paths = event.widget.tk.splitlist(event.data)
        folder_count = 0
        for file_path in file_paths:
            if not is_valid_folder(file_path):
                err_msg = ("Error! The last drop includes a folder from an invalid location. " +
                           f"Valid Location: {APPROVED_MUSIC_FOLDER}.")
                status_font.config(text=err_msg, fg="red")
                insert_into_error_log(ErrorType.INVALID_FOLDER, err_msg)  # not using the error_id returned
                return

            if os.path.isdir(file_path):
                process_folder(file_path, tree)
                folder_count += 1

        # Update the label with the folder count
        status_font.config(text=f"Folders processed in the last drop: {folder_count}", fg="white")
    except Exception as e:
        # Catch any unhandled errors, update label and log error
        status_font.config(text=f"Error! {e}. Operation could not be completed; " +
                                "some folders may have been processed prior to the error.",
                           fg="red"
                           )
        insert_into_error_log(ErrorType.PROCESSING_ERROR, f"{e}")


def process_folder(folder_path, tree):
    # No exception handling as this will be done within on_drop()
    m4p_files = [os.path.join(root, file) for root, _, files in os.walk(folder_path) for file in files if
                 file.endswith('.m4p')]
    for m4p_file in m4p_files:
        song_metadata = extract_metadata(m4p_file)
        is_duplicate = check_song_exists(song_metadata['song'], song_metadata['album'],
                                         song_metadata['artist'])
        insert_into_songs(is_duplicate, song_metadata)
        display_metadata(tree, song_metadata, is_duplicate)


def display_metadata(tree, metadata, is_duplicate):
    try:
        status = "Duplicate" if is_duplicate else "New"

        tree.insert("", "end", values=(
            metadata.get('song', 'N/A'), metadata.get('album', 'N/A'), metadata.get('artist', 'N/A'),
            metadata.get('approx_release_date', 'N/A'), status)
                    )

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.UNKNOWN_ERROR,
                                         f"Unexpected error in display_metadata: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred while displaying metadata. "
                             f"See Error Log - Error ID {error_id}."
                             )


# ---- Data Manipulation ----
def load_all_songs(tree):
    try:
        # Clear existing entries in the Treeview
        for row in tree.get_children():
            tree.delete(row)

        # Fetch all songs from the database
        songs = get_all_songs()

        # Insert each song into the Treeview
        for (song, album, artist, approx_release_date, time, file_size, created_date) in songs:
            tree.insert(
                "",
                "end",
                values=(song, album, artist, approx_release_date, time, file_size, created_date)
            )

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.DATABASE_ERROR,
                                         f"Unexpected error in load_all_songs: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred while loading songs from the database." +
                             f"See Error Log - Error ID {error_id}."
                             )


def load_processing_error_logs(tree):
    try:
        for row in tree.get_children():
            tree.delete(row)

        # Fetch all songs from the database
        processing_error_logs = get_processing_error_logs()

        # Insert each song into the Treeview
        for (err_id, error_type, error_message, created_date) in processing_error_logs:
            tree.insert(
                "",
                "end",
                values=(err_id, error_type, error_message, created_date)
            )

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.DATABASE_ERROR,
                                         f"Unexpected error in load_processing_error_logs: {e}"
                                         )
        messagebox.showerror("Error",
                             f"An unexpected error occurred while loading processing error logs. " +
                             f"See Error Log - Error ID {error_id}."
                             )


def load_all_error_logs(tree):
    try:
        for row in tree.get_children():
            tree.delete(row)

        # Fetch all songs from the database
        error_logs = get_all_error_logs()

        # Insert each song into the Treeview
        for (error_id, error_type, error_message, created_date) in error_logs:
            tree.insert(
                "",
                "end",
                values=(error_id, error_type, error_message, created_date)
            )

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.DATABASE_ERROR,
                                         f"Unexpected error in load_processing_error_logs: {e}"
                                         )
        messagebox.showerror("Error",
                             f"An unexpected error occurred while loading processing error logs. "
                             f"See Error Log - Error ID {error_id}."
                             )


def refresh_db_data(search_bar, tree):
    try:
        search_bar.delete(0, tk.END)  # clear search box
        load_all_songs(tree)  # reload all songs
        tree.focus_set()  # change focus to tree - so cursor not flashing in search box
        set_search_bar_placeholder(search_bar)  # reset the search bar placeholder

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.DATABASE_ERROR,
                                         f"Unexpected error in refresh_db_data: {e}")
        messagebox.showerror("Error",
                             f"An unexpected error occurred while refreshing the database data. "
                             f"See Error Log - Error ID {error_id}."
                             )


def refresh_err_data(tree, view_all):
    try:
        if view_all:
            load_all_error_logs(tree)
        else:
            load_processing_error_logs(tree)

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.DATABASE_ERROR,
                                         f"Unexpected error in refresh_err_data: {e}"
                                         )
        messagebox.showerror("Error",
                             f"An unexpected error occurred while refreshing the database data. "
                             f"See Error Log - Error ID {error_id}."
                             )


def search_song(tree, song_name, search_bar):
    try:
        # Clear current Treeview contents
        for row in tree.get_children():
            tree.delete(row)

        # Fetch songs that match the search query
        results = get_songs_by_name(song_name)

        # Insert matching songs into the Treeview
        for song in results:
            tree.insert("", "end", values=song)

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.DATABASE_ERROR,
                                         f"Unexpected error in search_song: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred while searching for songs. "
                             f"See Error Log - Error ID {error_id}."
                             )


def delete_selected_songs(full_delete, tree):
    try:
        selected_items = tree.selection()

        if not selected_items:
            messagebox.showwarning("No Selection", "Error: No song(s) have been selected.")
            return

        # Confirm deletion for all selected items
        if full_delete and not messagebox.askyesno("Confirm Delete",
                                                   "Are you sure you want to delete the selected song(s)?"
                                                   ):
            return

        # Loop through selected items and delete each from database and Treeview
        for selected_item in selected_items:
            item = tree.item(selected_item)  # retrieve song info for each selected item
            song_data = item['values']

            if full_delete:
                song, album, artist, *_ = song_data  # unpack relevant fields - '_' ignores unneeded fields
                delete_song(song, album, artist)  # delete from database

            # Remove from Treeview
            tree.delete(selected_item)

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.DATABASE_ERROR,
                                         f"Unexpected error in delete_selected_songs: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred while deleting selected songs. "
                             f"See Error Log - Error ID {error_id}."
                             )


def delete_selected_error_log(tree):
    try:
        selected_items = tree.selection()

        if not selected_items:
            messagebox.showwarning("No Selection", "Error: No error log(s) have been selected.")
            return

        # Confirm deletion for all selected items
        if not messagebox.askyesno("Confirm Delete",
                                   "Are you sure you want to delete the selected error log(s)?"):
            return

        # Loop through selected items and delete each from database and Treeview
        for selected_item in selected_items:
            item = tree.item(selected_item)  # retrieve error log info for each selected item
            error_log_data = item['values']

            error_id, error_message, created_date, *_ = error_log_data  # unpack rel fields, '_' ignores unneeded fields
            delete_error_log(error_id)  # delete from database

            # Remove from Treeview
            tree.delete(selected_item)

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.DATABASE_ERROR,
                                         f"Unexpected error in delete_selected_error_log: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred while deleting selected error logs. "
                             f"See Error Log - Error ID {error_id}."
                             )


# ---- Exporting ----
def export_to_excel(doc_prefix, col_headers, tree, hide_column_b):
    try:
        # Check if the Treeview is empty
        if not tree.get_children():
            messagebox.showwarning("Export Failed", "Error: No data to export.")
            return

        # Confirmation to export to Excel
        if not messagebox.askyesno("Confirm Excel Export", "Are you sure you to export this data to Excel?"):
            return

        # Save the exported Excel file to the path defined within the save_directory variable
        root_directory = os.path.expanduser("~")  # e.g. '/Users/nicholaspackham'

        # Using two different directories for testing and prod
        if IS_TEST_MODE:
            save_directory = os.path.join(root_directory, APP_ROOT_FOLDER_TEST_MODE)
        else:
            save_directory = os.path.join(root_directory, APP_ROOT_FOLDER)

        # Append 'Excel Exports' on to APP_ROOT_FOLDER*
        save_directory = os.path.join(save_directory, "Excel Exports")

        # Create the directory - if it does not exist
        if not os.path.exists(save_directory):
            os.makedirs(save_directory)

        # Generate the file name - which will provide the full file_path
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(save_directory, f"{doc_prefix}_{current_datetime}.xlsx")

        # Create a new Excel workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Songs"

        # Define styling for headers (excel)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Add headers with styling (excel)
        sheet.append(col_headers)
        for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(col_headers)):
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = thin_border

        # Add Treeview data to the Excel sheet
        for row_id in tree.get_children():
            row = tree.item(row_id)['values']
            sheet.append(row)

        # Add filters to the header row (excel)
        sheet.auto_filter.ref = sheet.dimensions

        # Auto-fit columns
        for col in sheet.columns:
            max_length = max(len(str(cell.value) or "") for cell in col)
            adjusted_width = max_length + 2
            sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        # ***Only relevant to Song trees (new and database) - not error log tree
        #   -Hide column "B" for the Song trees - Album column
        #   -Album column not really required, and gets in the way when using Apple Music and the Excel export together
        if hide_column_b:
            sheet.column_dimensions["B"].hidden = True

        # Save the workbook
        workbook.save(file_path)
        messagebox.showinfo(
            "Export Successful",
            "Success! Data has been exported to Excel.\n\n" +
            "Excel Export Location:\n" +
            f"{file_path}"
        )

        # Automatically open the file
        try:
            if os.name == 'nt':  # Windows
                os.startfile(file_path)
            elif os.name == 'posix':  # macOS and Linux
                subprocess.Popen(['open', file_path] if os.uname().sysname == 'Darwin' else ['xdg-open', file_path])
        except Exception as e:
            messagebox.showerror(
                "Error",
                "Error: Could not open file.\n\n" +
                "File Location:\n" +
                f"{e}"
            )

    except Exception as e:
        # Log the unexpected error and get the error ID
        error_id = insert_into_error_log(ErrorType.EXPORT_ERROR,
                                         f"Unexpected error in export_to_excel: {e}"
                                         )
        messagebox.showerror("Error",
                             "An unexpected error occurred during the Excel export process. "
                             f"See Error Log - Error ID {error_id}."
                             )


# ---- Backups ----
def save_database_backup():
    db_path = database_path()  # Ensure this returns the database path

    # Let the user select a directory
    folder_path = filedialog.askdirectory(
        title="Select Folder to Save Backup"
    )

    if folder_path:  # if a folder was selected
        file_path = os.path.join(folder_path, DATABASE_FILE_NAME)  # combine folder and database filename

        try:
            shutil.copy(db_path, file_path)  # copy the database to the chosen location
            messagebox.showinfo("Success", f"Database backed up to: {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to back up the database:\n{e}")
    else:
        messagebox.showinfo("Cancelled", "No folder selected. Backup cancelled.")
